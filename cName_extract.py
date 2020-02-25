# import shelve
import openpyxl
import os

wb = openpyxl.Workbook()

vals = ['cMean1', 'cMean2']

files = [f for f in os.listdir('./balances') if '.out' in f]
for fi in files:
    myFile = open('./balances/%s' % fi)
    content = myFile.read()
    lines = content.split('\n')
    wb.create_sheet(title=fi)
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == fi:
            break
    wb.active = s
    ws = wb.active
    counter = 1
    ws['A%s' % counter] = 'TIME'
    ws['B%s' % counter] = 'CMEAN'
    ws['C%s' % counter] = 'VALUE'
    # wb.save('text.xlsx')
    # Alternate between 1 and 2

    def alternate():
        alternate.x = not alternate.x
        return alternate.x

    alternate.x = True

    time = ''
    x = 1
    valid = False
    for line in lines:
        if line.lstrip().startswith(('Time', 'cMean')):
            valid = True
            print(fi, line)
        else:
            valid = False
        if valid:
            if line.lstrip().startswith(('Time')):
                time = line.split()[-1]

            if line.lstrip().startswith('cMean'):
                cMeans = line.split()
                tp = vals[alternate()]
                for cmean in cMeans[4:]:
                    counter = counter + 1
                    ws['A%s' % counter] = time
                    ws['B%s' % counter] = tp
                    ws['C%s' % counter] = cmean
                    # print(time, '\t', tp, '\t', cmean)
    wb.save('Final.xlsx')
